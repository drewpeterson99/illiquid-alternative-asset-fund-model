"""Annual roll-up from monthly projections (ModelTemplate I17:Q37)."""

from __future__ import annotations

from datetime import date

import openpyxl
import pandas as pd
import pytest

from fund_engine import (
    DEFAULT_AS_OF_DATE,
    _FUND_ANNUAL_GROUP_KEYS,
    _PORTFOLIO_ANNUAL_TAG_CATEGORIES,
    Portfolio,
    aggregate_monthly_to_annual,
    aggregate_portfolio_annual_tagged,
    build_fund_detail,
    build_portfolio_monthly_detail,
)
from tests.factories import make_fund
from tests.model_template import PROJECTION_START, TEMPLATE_PATH

_EXCEL_ANNUAL_ROWS: dict[int, str] = {
    18: "total_commitment",
    19: "effective_commitment",
    21: "legal_unfunded",
    22: "remaining_effective_unfunded",
    24: "beginning_nav",
    25: "capital_called",
    26: "distributions",
    27: "net_income",
    28: "ending_nav",
    32: "roc",
    33: "dividend",
    34: "gain_on_sale",
    37: "average_nav",
    38: "period_return",
    39: "dividend_return",
}


def _load_excel_annual_golden() -> list[tuple[date, dict[str, float]]]:
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    ws = wb["Model"]
    units = float(ws["D1"].value)
    records: list[tuple[date, dict[str, float]]] = []
    for col in range(10, 17):
        year_end = ws.cell(17, col).value
        if year_end is None:
            continue
        if hasattr(year_end, "date"):
            year_end = year_end.date()
        metrics: dict[str, float] = {}
        for row in _EXCEL_ANNUAL_ROWS:
            raw = ws.cell(row, col).value
            if raw in (None, "-", "n/m"):
                continue
            metrics[_EXCEL_ANNUAL_ROWS[row]] = float(raw) * (
                units if row not in (38, 39) else 1.0
            )
        records.append((year_end, metrics))
    return records


@pytest.fixture(scope="module")
def excel_annual_golden() -> list[tuple[date, dict[str, float]]]:
    return _load_excel_annual_golden()


def test_annual_matches_model_template(
    template_fund,
    engine_projection,
    excel_annual_golden: list[tuple[date, dict[str, float]]],
) -> None:
    annual = aggregate_monthly_to_annual(
        engine_projection, as_of_date=PROJECTION_START
    )
    monthly_end_dates = set(pd.to_datetime(engine_projection["end_date"]).dt.date)
    for year_end, expected in excel_annual_golden:
        if year_end not in monthly_end_dates:
            continue
        match = annual.loc[annual["year"] == year_end.year]
        if match.empty:
            continue
        row = match.iloc[0]
        for metric, value in expected.items():
            assert row[metric] == pytest.approx(value), metric


def test_funds_annual_output(assumptions_funds) -> None:
    funds_monthly = build_fund_detail(assumptions_funds, DEFAULT_AS_OF_DATE)
    funds_annual = aggregate_monthly_to_annual(
        funds_monthly,
        as_of_date=DEFAULT_AS_OF_DATE,
        group_keys=_FUND_ANNUAL_GROUP_KEYS,
    )
    assert set(funds_annual["fund"]).issubset({f.name for f in assumptions_funds})
    assert len(funds_annual) > 0

    for _, arow in funds_annual.iterrows():
        operating = funds_monthly[
            (funds_monthly["fund"] == arow["fund"])
            & (funds_monthly["period"] > 0)
            & (pd.to_datetime(funds_monthly["end_date"]).dt.year == arow["year"])
        ]
        assert arow["operating_months"] == len(operating)



def test_annual_rollup_rules(template_fund) -> None:
    """Period 0 is reference-only; returns and flows follow annual formulas."""
    monthly = template_fund.project(PROJECTION_START, date(2027, 12, 31))
    annual = aggregate_monthly_to_annual(monthly, as_of_date=PROJECTION_START)
    row = annual.loc[annual["year"] == 2026].iloc[0]

    operating_2026 = monthly[
        (monthly["period"] > 0)
        & (pd.to_datetime(monthly["end_date"]).dt.year == 2026)
    ]
    assert row["capital_called"] == pytest.approx(operating_2026["capital_called"].sum())
    assert row["beginning_nav"] == pytest.approx(
        monthly.loc[monthly["period"] == 0, "nav"].iloc[0]
    )
    assert row["operating_months"] == len(operating_2026)
    assert row["period_return"] == pytest.approx(row["net_income"] / row["average_nav"])
    assert row["dividend_return"] == pytest.approx(abs(row["dividend"]) / row["average_nav"])


def test_mid_year_termination_included_in_annual() -> None:
    fund = make_fund(
        name="Mid Year Exit",
        funded_amount=60_000_000,
        stated_nav=72_000_000,
        invest_end_date=date(2024, 1, 31),
        reinvest_end_date=date(2025, 6, 30),
        termination_date=date(2030, 6, 30),
    )
    as_of = date(2025, 12, 31)
    monthly = fund.project(as_of, fund.termination_date)
    row = aggregate_monthly_to_annual(monthly, as_of_date=as_of)
    row = row.loc[row["year"] == 2030].iloc[0]

    assert row["year"] == 2030
    assert row["operating_months"] == 6
    assert row["ending_nav"] == pytest.approx(
        monthly.loc[monthly["end_date"] == fund.termination_date, "nav"].iloc[-1]
    )


def test_portfolio_annual_tagged(assumptions_funds) -> None:
    portfolio_monthly = Portfolio(assumptions_funds).aggregate(DEFAULT_AS_OF_DATE)
    portfolio_annual = aggregate_portfolio_annual_tagged(
        build_portfolio_monthly_detail(assumptions_funds, DEFAULT_AS_OF_DATE),
        as_of_date=DEFAULT_AS_OF_DATE,
    )
    assert (portfolio_annual["currency"] == "USD").all()

    total = portfolio_annual[
        (portfolio_annual["tag_category"] == "TOTAL")
        & (portfolio_annual["tag"] == "TOTAL")
    ]
    assert not total.empty
    assert set(portfolio_annual["tag_category"]) >= set(_PORTFOLIO_ANNUAL_TAG_CATEGORIES)

    plain = aggregate_monthly_to_annual(
        portfolio_monthly, as_of_date=DEFAULT_AS_OF_DATE, group_keys=("currency",)
    )
    merged = total.merge(
        plain, on=["currency", "year", "operating_months"], suffixes=("_tag", "")
    )
    assert merged["ending_nav_tag"].equals(merged["ending_nav"])
